import os, io, re, base64
from datetime import datetime, date, timezone
from typing import Optional, Dict, Any, List, Tuple

import requests
from msal import ConfidentialClientApplication
from openpyxl import load_workbook


# ==================== CONFIG ====================
GRAPH = "https://graph.microsoft.com/v1.0"

MAILBOX_UPN = "apatil@glassdoctordfw.com"
DRIVE_ID = os.getenv("DRIVE_ID", "")
ITEM_ID  = os.getenv("ITEM_ID", "")

SHEET_AUTO        = os.getenv("SHEET_AUTO", "Auto")
SHEET_RETAIL      = os.getenv("SHEET_RETAIL", "Retail")
SHEET_BUILDER     = os.getenv("SHEET_BUILDER", "Builder")
SHEET_COMMERCIAL  = os.getenv("SHEET_COMMERCIAL", "Commercial")
SHEET_MULTIFAMILY = os.getenv("SHEET_MULTIFAMILY", "MultiFamily")

# -------- Subjects --------
# Auto
SUBJECT_AUTO_BOOKED       = "Daily Metrics Report - Auto Glass Booked"
SUBJECT_AUTO_OPP_REPORT   = "Daily Metrics Report - Opportunity Report- Auto"

# Retail
SUBJECT_PREV_BOOKED_WO    = "Daily Metrics Report - Previous Booked Work Orders"
SUBJECT_DAILY_HUDDLE      = "Daily Metrics Report - Daily Huddle"
SUBJECT_SOLD_BY_EMP       = "Daily Metrics Report - Monthly Sold Jobs by Employee"
SUBJECT_SOLD_TO_CONVERT   = "Daily Metrics Report - Sold Estimates to be converted- Retail"
SUBJECT_UNSCHEDULED       = "Daily Metrics Report - Unscheduled Jobs"
SUBJECT_RETAIL_OPP_REPORT = "Daily Metrics Report - Opportunity Report- Retail"

# Builder/Commercial/MultiFamily
SUBJECT_BUILDER_WO_BOOKED = "Daily Metrics Report - Work Orders Booked - Builder"
SUBJECT_BUILDER_SOLD_JOBS = "Daily Metrics Report - Sold Jobs - Builder"

SUBJECT_COMM_WO_BOOKED    = "Daily Metrics Report - Work Orders Booked - Commercial"
SUBJECT_COMM_SOLD_JOBS    = "Daily Metrics Report - Sold Jobs - Commercial"

SUBJECT_MF_WO_BOOKED      = "Daily Metrics Report - Work Orders Booked - MultiFamily"
SUBJECT_MF_SOLD_JOBS      = "Daily Metrics Report - Sold Jobs - MultiFamily"

SUBJECT_OPP_EST_FOLLOWUP  = "Daily Metrics Report - Opportunity and Estimate Follow Up"

RETAIL_NAMES = {"denton", "dallas", "carrollton", "arlington", "colleyville"}

HUDDLE_BUILDER_NAME = "SpecOps Builder Division"
HUDDLE_COMM_NAME    = "SpecOps Commercial Division"
HUDDLE_MF_NAME      = "SpecOps Multifamily Division"  # matches your Daily Huddle naming

FOLLOWUP_BUILDER_BU = "SpecOps Builder Division"
FOLLOWUP_COMM_BU    = "SpecOps Commercial Division"
FOLLOWUP_MF_BU      = "SpecOps MultiFamily Division"  # note: you spelled MultiFamily differently here (kept as you wrote)


# ==================== Sheet column mappings (1-based) ====================
# AUTO: Date | Day | Auto Glass Booked | Total Jobs Booked | Auto Closing Rate | Auto Glass Installed
A_COL_DATE = 1
A_COL_AUTO_GLASS_BOOKED = 3
A_COL_TOTAL_JOBS_BOOKED = 4
A_COL_AUTO_CLOSING_RATE = 5
A_COL_AUTO_GLASS_INSTALLED = 6

# RETAIL (10):
# Date | Day | Flat Opp Booked | Flat Opp Phone Sales | Closed Avg Sale | Close Rate Field | Sold By | Total Rev Installed | Sold Jobs to Convert | Unscheduled Jobs
R_COL_DATE = 1
R_COL_FLAT_OPP_BOOKED     = 3
R_COL_PHONE_SALES         = 4
R_COL_CLOSED_AVG_SALE     = 5
R_COL_CLOSE_RATE_FIELD    = 6
R_COL_SOLD_BY_REPORT      = 7
R_COL_TOTAL_REV_INST      = 8
R_COL_SOLD_TO_CONVERT     = 9
R_COL_UNSCHEDULED         = 10

# BUILDER/COMM/MF (7):
# Date | Day | Work Orders Booked | Opportunities/Estimates | Sold Jobs | Total Revenue | Close Rate Field
X_COL_DATE = 1
X_COL_WO_BOOKED  = 3
X_COL_OPP_EST    = 4
X_COL_SOLD_JOBS  = 5
X_COL_TOTAL_REV  = 6
X_COL_CLOSE_RATE = 7


# ==================== Helpers ====================
def must_env(name: str) -> str:
    v = os.getenv(name)
    if not v:
        raise RuntimeError(f"Missing environment variable: {name}")
    return v

def parse_dt(dt_str: str) -> datetime:
    if not dt_str:
        return datetime(1970, 1, 1, tzinfo=timezone.utc)
    if dt_str.endswith("Z"):
        dt_str = dt_str.replace("Z", "+00:00")
    return datetime.fromisoformat(dt_str)

def try_parse_any_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def normalize_header(x) -> str:
    return str(x).strip().lower()

def find_col_idx(header_row: List[Any], targets_lower: set) -> Optional[int]:
    header = [normalize_header(h) for h in header_row]
    for i, h in enumerate(header):
        if h in targets_lower:
            return i
    for i, h in enumerate(header):
        for t in targets_lower:
            if t in h:
                return i
    return None

def last_non_empty_in_col(body_rows: List[List[Any]], col_idx: int):
    for r in reversed(body_rows):
        if len(r) <= col_idx:
            continue
        v = r[col_idx]
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return v
    return None

def parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    s = s.replace(",", "").replace("$", "").replace(" ", "")
    if s.endswith("%"):
        s = s[:-1]
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def count_non_null(body_rows: List[List[Any]], col_idx: int) -> int:
    c = 0
    for r in body_rows:
        if len(r) <= col_idx:
            continue
        v = r[col_idx]
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        c += 1
    return c

def extract_date_from_filename(fname: str) -> date:
    if not fname:
        raise ValueError("Missing filename for date extraction.")
    m = re.search(r"(\d{1,4})[._/\-](\d{1,2})[._/\-](\d{1,4})", fname)
    if not m:
        raise ValueError(f"Could not extract date from filename: {fname}")
    a, b, c = m.groups()
    nums = list(map(int, (a, b, c)))
    if nums[0] > 31:
        yyyy, mm, dd = nums[0], nums[1], nums[2]
    else:
        mm, dd, yy = nums
        yyyy = 2000 + yy if yy < 100 else yy
    return date(yyyy, mm, dd)

def read_xlsx_first_sheet_rows(xlsx_bytes: bytes) -> List[List[Any]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else v) for v in r])
    return rows


# ==================== Auth (App-only) ====================
def get_token() -> str:
    tenant_id = must_env("TENANT_ID")
    client_id = must_env("CLIENT_ID")
    client_secret = must_env("CLIENT_SECRET")

    app = ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Token error: {result.get('error')} - {result.get('error_description')}")
    return result["access_token"]

def graph_get(token: str, url: str, params: Optional[dict] = None) -> Dict[str, Any]:
    headers = {"Authorization": f"Bearer {token}", "ConsistencyLevel": "eventual"}
    r = requests.get(url, headers=headers, params=params, timeout=60)
    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()
    return r.json()

def graph_get_bytes(token: str, url: str, params: Optional[dict] = None) -> bytes:
    headers = {"Authorization": f"Bearer {token}", "ConsistencyLevel": "eventual"}
    r = requests.get(url, headers=headers, params=params, timeout=120)
    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()
    return r.content

def graph_post(token: str, url: str, payload: dict) -> dict:
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    r = requests.post(url, headers=headers, json=payload, timeout=60)
    if not r.ok:
        try:
            print("Graph POST error payload:", r.json())
        except Exception:
            print("Graph POST error text:", r.text)
        r.raise_for_status()
    return r.json()


# ==================== Outlook helpers ====================
def latest_message_for_subject(token: str, mailbox_upn: str, subject_phrase: str) -> Optional[Dict[str, Any]]:
    url = f"{GRAPH}/users/{mailbox_upn}/mailFolders/Inbox/messages"
    params = {
        "$select": "id,subject,receivedDateTime,from,hasAttachments",
        "$top": "25",
        "$search": f"\"{subject_phrase}\"",
    }
    data = graph_get(token, url, params=params)
    msgs = data.get("value", [])
    phrase = subject_phrase.lower()
    candidates = [m for m in msgs if phrase in (m.get("subject") or "").lower()]
    if not candidates:
        return None
    candidates.sort(key=lambda m: parse_dt(m.get("receivedDateTime", "")), reverse=True)
    return candidates[0]

def get_first_xlsx_attachment_from_message(token: str, mailbox_upn: str, message_id: str) -> Tuple[Optional[str], Optional[bytes]]:
    url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments"
    data = graph_get(token, url, params={"$top": "50"})
    atts = data.get("value", [])
    for a in atts:
        name = (a.get("name") or "")
        if name.lower().endswith(".xlsx"):
            cb = a.get("contentBytes")
            if cb:
                return name, base64.b64decode(cb)
            att_id = a.get("id")
            if att_id:
                raw_url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments/{att_id}/$value"
                b = graph_get_bytes(token, raw_url)
                return name, b
    return None, None


# ==================== SharePoint download/upload ====================
def download_sharepoint_excel(token: str, drive_id: str, item_id: str) -> bytes:
    url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
    return graph_get_bytes(token, url)

def upload_sharepoint_excel_resumable(token: str, drive_id: str, item_id: str, content: bytes) -> None:
    create_url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/createUploadSession"
    payload = {"item": {"@microsoft.graph.conflictBehavior": "replace"}}
    sess = graph_post(token, create_url, payload)
    upload_url = sess["uploadUrl"]

    chunk_size = 10 * 1024 * 1024
    total = len(content)
    start = 0

    while start < total:
        end = min(start + chunk_size, total) - 1
        chunk = content[start:end + 1]

        headers = {
            "Content-Length": str(len(chunk)),
            "Content-Range": f"bytes {start}-{end}/{total}",
        }
        r = requests.put(upload_url, headers=headers, data=chunk, timeout=180)

        if r.status_code in (200, 201):
            return
        if r.status_code == 202:
            start = end + 1
            continue

        try:
            print("Upload session error payload:", r.json())
        except Exception:
            print("Upload session error text:", r.text)
        r.raise_for_status()

def upload_sharepoint_excel(token: str, drive_id: str, item_id: str, content: bytes) -> None:
    url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    r = requests.put(url, headers=headers, data=content, timeout=180)

    if r.status_code == 423:
        print("File is locked (423). Falling back to upload session...")
        upload_sharepoint_excel_resumable(token, drive_id, item_id, content)
        return

    if not r.ok:
        try:
            print("Simple upload error payload:", r.json())
        except Exception:
            print("Simple upload error text:", r.text)
        r.raise_for_status()


# ==================== Workbook helpers ====================
def build_sheet_date_row_map(ws, date_col: int) -> Dict[date, int]:
    mapping = {}
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=date_col).value
        d = try_parse_any_date(v)
        if d:
            mapping[d] = row
    return mapping

def get_row_value_by_name(rows: List[List[Any]], name_exact: str, value_col_targets: set) -> float:
    header, body = rows[0], rows[1:]
    name_col = find_col_idx(header, {"name"})
    val_col  = find_col_idx(header, value_col_targets)
    if name_col is None or val_col is None:
        raise ValueError(f"[DAILY HUDDLE] Missing required columns for {value_col_targets}. Header: {header}")
    for r in body:
        if len(r) <= max(name_col, val_col):
            continue
        if str(r[name_col]).strip().lower() == name_exact.strip().lower():
            return parse_number(r[val_col])
    raise ValueError(f"[DAILY HUDDLE] Name '{name_exact}' not found for {value_col_targets}.")


# ==================== Parsing functions ====================
def parse_auto_booked_attachment(rows: List[List[Any]]) -> Tuple[Any, Any]:
    header, body = rows[0], rows[1:]
    job_col = find_col_idx(header, {"job #", "job#", "job number", "job no"})
    sub_col = find_col_idx(header, {"jobs subtotal", "job subtotal", "subtotal"})
    if job_col is None or sub_col is None:
        raise ValueError(f"[AUTO BOOKED] Missing required columns. Header: {header}")
    return last_non_empty_in_col(body, job_col), last_non_empty_in_col(body, sub_col)

def parse_work_orders_booked(rows: List[List[Any]]) -> Any:
    header, body = rows[0], rows[1:]
    job_col = find_col_idx(header, {"job #", "job#", "job number", "job no"})
    if job_col is None:
        raise ValueError(f"[WO BOOKED] Missing 'Job #' column. Header: {header}")
    return last_non_empty_in_col(body, job_col)

def parse_sold_jobs_last_estimate_id(rows: List[List[Any]]) -> Any:
    header, body = rows[0], rows[1:]
    est_col = find_col_idx(header, {"estimate id", "estimateid"})
    if est_col is None:
        raise ValueError(f"[SOLD JOBS] Missing 'Estimate Id' column. Header: {header}")
    return last_non_empty_in_col(body, est_col)

def parse_prev_booked_work_orders(rows: List[List[Any]]) -> Any:
    header, body = rows[0], rows[1:]
    job_col = find_col_idx(header, {"job #", "job#", "job number", "job no"})
    if job_col is None:
        raise ValueError(f"[RETAIL PREV BOOKED WO] Missing 'Job #' column. Header: {header}")
    return last_non_empty_in_col(body, job_col)

def parse_sold_by_report(rows: List[List[Any]]) -> Any:
    return parse_sold_jobs_last_estimate_id(rows)

def parse_sold_to_convert(rows: List[List[Any]]) -> int:
    header, body = rows[0], rows[1:]
    num_col = find_col_idx(header, {"number"})
    if num_col is None:
        raise ValueError(f"[RETAIL SOLD TO CONVERT] Missing 'Number' column. Header: {header}")
    return count_non_null(body, num_col)

def parse_unscheduled_jobs(rows: List[List[Any]]) -> Any:
    header, body = rows[0], rows[1:]
    job_col = find_col_idx(header, {"job #", "job#", "job number", "job no"})
    if job_col is None:
        raise ValueError(f"[RETAIL UNSCHEDULED] Missing 'Job #' column. Header: {header}")
    return last_non_empty_in_col(body, job_col)

def parse_retail_metrics_from_huddle(rows: List[List[Any]]) -> Tuple[float, float, float]:
    """
    Closed Avg Sale = AVERAGE across Retail BU names
    Close Rate Field = AVERAGE across Retail BU names
    Total Rev Installed = SUM Completed Revenue across Retail BU names
    """
    header, body = rows[0], rows[1:]
    name_col = find_col_idx(header, {"name"})
    cas_col  = find_col_idx(header, {"closed average sale"})
    cr_col   = find_col_idx(header, {"close rate"})
    rev_col  = find_col_idx(header, {"completed revenue"})
    if name_col is None or cas_col is None or cr_col is None or rev_col is None:
        raise ValueError(f"[DAILY HUDDLE] Missing columns for Retail metrics. Header: {header}")

    cas_sum = cas_cnt = 0
    cr_sum  = cr_cnt  = 0
    rev_sum = 0.0

    for r in body:
        if len(r) <= max(name_col, cas_col, cr_col, rev_col):
            continue
        nm = str(r[name_col]).strip().lower()
        if nm not in RETAIL_NAMES:
            continue
        cas_sum += parse_number(r[cas_col]); cas_cnt += 1
        cr_sum  += parse_number(r[cr_col]);  cr_cnt  += 1
        rev_sum += parse_number(r[rev_col])

    cas_avg = (cas_sum / cas_cnt) if cas_cnt else 0.0
    cr_avg  = (cr_sum  / cr_cnt)  if cr_cnt  else 0.0
    return round(cas_avg, 2), round(cr_avg, 2), round(rev_sum, 2)

def parse_auto_closing_rate(rows: List[List[Any]]) -> float:
    """
    Auto Closing Rate = (# rows with Estimate Status == 'Sold') / (total non-empty Estimate Status rows)
    """
    header, body = rows[0], rows[1:]
    status_col = find_col_idx(header, {"estimate status"})
    if status_col is None:
        raise ValueError(f"[AUTO OPP REPORT] Missing 'Estimate Status' column. Header: {header}")

    total = 0
    sold = 0
    for r in body:
        if len(r) <= status_col:
            continue
        v = r[status_col]
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        total += 1
        if str(v).strip().lower() == "sold":
            sold += 1

    return round((sold / total) if total else 0.0, 4)

def parse_retail_phone_sales_sold_count(rows: List[List[Any]]) -> int:
    """
    Flat Opportunities Booked Phone Sales = count rows with Estimate Status == 'Sold'
    """
    header, body = rows[0], rows[1:]
    status_col = find_col_idx(header, {"estimate status"})
    if status_col is None:
        raise ValueError(f"[RETAIL OPP REPORT] Missing 'Estimate Status' column. Header: {header}")

    sold = 0
    for r in body:
        if len(r) <= status_col:
            continue
        v = r[status_col]
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if str(v).strip().lower() == "sold":
            sold += 1
    return sold

def parse_followup_counts_by_bu(rows: List[List[Any]], bu_name: str) -> int:
    """
    Opportunity and Estimate Follow Up:
    count rows where Business Unit == bu_name
    """
    header, body = rows[0], rows[1:]
    bu_col = find_col_idx(header, {"business unit"})
    if bu_col is None:
        raise ValueError(f"[FOLLOW UP] Missing 'Business Unit' column. Header: {header}")

    cnt = 0
    for r in body:
        if len(r) <= bu_col:
            continue
        v = r[bu_col]
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if str(v).strip().lower() == bu_name.strip().lower():
            cnt += 1
    return cnt


# ==================== Apply functions ====================
def apply_auto_updates(wb, file_date: date,
                       auto_glass_booked=None,
                       total_jobs_booked=None,
                       auto_closing_rate=None,
                       auto_glass_installed=None) -> int:
    ws = wb[SHEET_AUTO]
    row = build_sheet_date_row_map(ws, A_COL_DATE).get(file_date)
    if not row:
        raise RuntimeError(f"[AUTO] Date {file_date} not found in '{SHEET_AUTO}'.")
    u = 0
    if auto_glass_booked is not None:
        ws.cell(row=row, column=A_COL_AUTO_GLASS_BOOKED).value = auto_glass_booked; u += 1
    if total_jobs_booked is not None:
        ws.cell(row=row, column=A_COL_TOTAL_JOBS_BOOKED).value = total_jobs_booked; u += 1
    if auto_closing_rate is not None:
        ws.cell(row=row, column=A_COL_AUTO_CLOSING_RATE).value = auto_closing_rate; u += 1
    if auto_glass_installed is not None:
        ws.cell(row=row, column=A_COL_AUTO_GLASS_INSTALLED).value = auto_glass_installed; u += 1
    return u

def apply_retail_updates(wb, file_date: date,
                         flat_opp_booked=None,
                         phone_sales=None,
                         closed_avg_sale=None,
                         close_rate_field=None,
                         sold_by_report=None,
                         total_rev_installed=None,
                         sold_jobs_to_convert=None,
                         unscheduled_jobs=None) -> int:
    ws = wb[SHEET_RETAIL]
    row = build_sheet_date_row_map(ws, R_COL_DATE).get(file_date)
    if not row:
        raise RuntimeError(f"[RETAIL] Date {file_date} not found in '{SHEET_RETAIL}'.")
    u = 0
    if flat_opp_booked is not None:
        ws.cell(row=row, column=R_COL_FLAT_OPP_BOOKED).value = flat_opp_booked; u += 1
    if phone_sales is not None:
        ws.cell(row=row, column=R_COL_PHONE_SALES).value = phone_sales; u += 1
    if closed_avg_sale is not None:
        ws.cell(row=row, column=R_COL_CLOSED_AVG_SALE).value = closed_avg_sale; u += 1
    if close_rate_field is not None:
        ws.cell(row=row, column=R_COL_CLOSE_RATE_FIELD).value = close_rate_field; u += 1
    if sold_by_report is not None:
        ws.cell(row=row, column=R_COL_SOLD_BY_REPORT).value = sold_by_report; u += 1
    if total_rev_installed is not None:
        ws.cell(row=row, column=R_COL_TOTAL_REV_INST).value = total_rev_installed; u += 1
    if sold_jobs_to_convert is not None:
        ws.cell(row=row, column=R_COL_SOLD_TO_CONVERT).value = sold_jobs_to_convert; u += 1
    if unscheduled_jobs is not None:
        ws.cell(row=row, column=R_COL_UNSCHEDULED).value = unscheduled_jobs; u += 1
    return u

def apply_7col_updates(wb, sheet_name: str, file_date: date,
                       wo_booked=None,
                       opp_est=None,
                       sold_jobs=None,
                       total_rev=None,
                       close_rate=None) -> int:
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found. Found: {wb.sheetnames}")
    ws = wb[sheet_name]
    row = build_sheet_date_row_map(ws, X_COL_DATE).get(file_date)
    if not row:
        raise RuntimeError(f"[{sheet_name}] Date {file_date} not found in '{sheet_name}'.")
    u = 0
    if wo_booked is not None:
        ws.cell(row=row, column=X_COL_WO_BOOKED).value = wo_booked; u += 1
    if opp_est is not None:
        ws.cell(row=row, column=X_COL_OPP_EST).value = opp_est; u += 1
    if sold_jobs is not None:
        ws.cell(row=row, column=X_COL_SOLD_JOBS).value = sold_jobs; u += 1
    if total_rev is not None:
        ws.cell(row=row, column=X_COL_TOTAL_REV).value = total_rev; u += 1
    if close_rate is not None:
        ws.cell(row=row, column=X_COL_CLOSE_RATE).value = close_rate; u += 1
    return u


# ==================== MAIN ====================
def main():
    if not DRIVE_ID or not ITEM_ID:
        raise RuntimeError("Missing DRIVE_ID or ITEM_ID environment variables.")

    token = get_token()

    # ---- Latest messages ----
    msg_auto_booked     = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_AUTO_BOOKED)
    msg_auto_opp        = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_AUTO_OPP_REPORT)

    msg_huddle          = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_DAILY_HUDDLE)

    msg_prev_booked     = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_PREV_BOOKED_WO)
    msg_sold_by         = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_SOLD_BY_EMP)
    msg_sold_conv       = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_SOLD_TO_CONVERT)
    msg_unsched         = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_UNSCHEDULED)
    msg_retail_opp      = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_RETAIL_OPP_REPORT)

    msg_builder_wo      = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_BUILDER_WO_BOOKED)
    msg_builder_sold    = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_BUILDER_SOLD_JOBS)

    msg_comm_wo         = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_COMM_WO_BOOKED)
    msg_comm_sold       = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_COMM_SOLD_JOBS)

    msg_mf_wo           = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_MF_WO_BOOKED)
    msg_mf_sold         = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_MF_SOLD_JOBS)

    msg_followup        = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_OPP_EST_FOLLOWUP)

    if not any([msg_auto_booked, msg_auto_opp, msg_huddle,
                msg_prev_booked, msg_sold_by, msg_sold_conv, msg_unsched, msg_retail_opp,
                msg_builder_wo, msg_builder_sold, msg_comm_wo, msg_comm_sold, msg_mf_wo, msg_mf_sold,
                msg_followup]):
        print("No matching emails found.")
        return

    # ---- Download workbook once ----
    xls_bytes = download_sharepoint_excel(token, DRIVE_ID, ITEM_ID)
    wb = load_workbook(io.BytesIO(xls_bytes))

    total_updates = 0

    # ================= AUTO =================
    if msg_auto_booked:
        mid = msg_auto_booked["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            d = extract_date_from_filename(fname)
            rows = read_xlsx_first_sheet_rows(content)
            a_booked, a_total = parse_auto_booked_attachment(rows)
            n = apply_auto_updates(wb, d, auto_glass_booked=a_booked, total_jobs_booked=a_total)
            total_updates += n
            print(f"[Auto Booked] {fname} | file_date={d} | updated_cells={n}")

    if msg_auto_opp:
        mid = msg_auto_opp["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            d = extract_date_from_filename(fname)
            rows = read_xlsx_first_sheet_rows(content)
            closing = parse_auto_closing_rate(rows)
            n = apply_auto_updates(wb, d, auto_closing_rate=closing)
            total_updates += n
            print(f"[Auto Closing Rate] {fname} | file_date={d} | closing_rate={closing} | updated_cells={n}")

    # ================= DAILY HUDDLE (Auto + Retail + Builder/Comm/MF rev & close rate) =================
    if msg_huddle:
        mid = msg_huddle["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            d = extract_date_from_filename(fname)
            rows = read_xlsx_first_sheet_rows(content)

            auto_installed = get_row_value_by_name(rows, "Auto Glass", {"completed revenue"})
            n_auto = apply_auto_updates(wb, d, auto_glass_installed=auto_installed)

            r_cas, r_cr, r_rev = parse_retail_metrics_from_huddle(rows)
            n_retail = apply_retail_updates(wb, d, closed_avg_sale=r_cas, close_rate_field=r_cr, total_rev_installed=r_rev)

            b_rev = get_row_value_by_name(rows, HUDDLE_BUILDER_NAME, {"completed revenue"})
            b_cr  = get_row_value_by_name(rows, HUDDLE_BUILDER_NAME, {"close rate"})
            n_builder = apply_7col_updates(wb, SHEET_BUILDER, d, total_rev=b_rev, close_rate=b_cr)

            c_rev = get_row_value_by_name(rows, HUDDLE_COMM_NAME, {"completed revenue"})
            c_cr  = get_row_value_by_name(rows, HUDDLE_COMM_NAME, {"close rate"})
            n_comm = apply_7col_updates(wb, SHEET_COMMERCIAL, d, total_rev=c_rev, close_rate=c_cr)

            m_rev = get_row_value_by_name(rows, HUDDLE_MF_NAME, {"completed revenue"})
            m_cr  = get_row_value_by_name(rows, HUDDLE_MF_NAME, {"close rate"})
            n_mf = apply_7col_updates(wb, SHEET_MULTIFAMILY, d, total_rev=m_rev, close_rate=m_cr)

            total_updates += (n_auto + n_retail + n_builder + n_comm + n_mf)
            print(f"[Daily Huddle] {fname} | file_date={d} | auto={n_auto} retail={n_retail} builder={n_builder} comm={n_comm} mf={n_mf}")

    # ================= RETAIL =================
    if msg_prev_booked:
        mid = msg_prev_booked["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            d = extract_date_from_filename(fname)
            rows = read_xlsx_first_sheet_rows(content)
            flat_opp = parse_prev_booked_work_orders(rows)
            n = apply_retail_updates(wb, d, flat_opp_booked=flat_opp)
            total_updates += n
            print(f"[Retail Prev Booked WO] {fname} | file_date={d} | updated_cells={n}")

    sold_by_date = None
    if msg_sold_by:
        mid = msg_sold_by["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            sold_by_date = extract_date_from_filename(fname)
            rows = read_xlsx_first_sheet_rows(content)
            sold_by_val = parse_sold_by_report(rows)
            n = apply_retail_updates(wb, sold_by_date, sold_by_report=sold_by_val)
            total_updates += n
            print(f"[Retail Sold By] {fname} | file_date={sold_by_date} | updated_cells={n}")

    if msg_sold_conv:
        mid = msg_sold_conv["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            if not sold_by_date:
                raise RuntimeError("Sold-to-convert needs Sold By Report date, but Sold By Report was not found/parsed.")
            rows = read_xlsx_first_sheet_rows(content)
            count_val = parse_sold_to_convert(rows)
            n = apply_retail_updates(wb, sold_by_date, sold_jobs_to_convert=count_val)
            total_updates += n
            print(f"[Retail Sold->Convert] {fname} | file_date(using SoldBy)={sold_by_date} | updated_cells={n}")

    if msg_unsched:
        mid = msg_unsched["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            if not sold_by_date:
                raise RuntimeError("Unscheduled Jobs needs Sold By Report date, but Sold By Report was not found/parsed.")
            rows = read_xlsx_first_sheet_rows(content)
            unsched_val = parse_unscheduled_jobs(rows)
            n = apply_retail_updates(wb, sold_by_date, unscheduled_jobs=unsched_val)
            total_updates += n
            print(f"[Retail Unscheduled] {fname} | file_date(using SoldBy)={sold_by_date} | updated_cells={n}")

    if msg_retail_opp:
        mid = msg_retail_opp["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            d = extract_date_from_filename(fname)
            rows = read_xlsx_first_sheet_rows(content)
            phone_sales_sold = parse_retail_phone_sales_sold_count(rows)
            n = apply_retail_updates(wb, d, phone_sales=phone_sales_sold)
            total_updates += n
            print(f"[Retail Phone Sales Sold] {fname} | file_date={d} | sold_count={phone_sales_sold} | updated_cells={n}")

    # ================= Builder/Comm/MF: WO Booked + Sold Jobs =================
    def handle_wo_and_sold(sheet_name: str, wo_msg, sold_msg, label: str):
        nonlocal total_updates
        if wo_msg:
            mid = wo_msg["id"]
            fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
            if content:
                d = extract_date_from_filename(fname)
                rows = read_xlsx_first_sheet_rows(content)
                wo_val = parse_work_orders_booked(rows)
                n = apply_7col_updates(wb, sheet_name, d, wo_booked=wo_val)
                total_updates += n
                print(f"[{label} WO Booked] {fname} | file_date={d} | updated_cells={n}")
        if sold_msg:
            mid = sold_msg["id"]
            fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
            if content:
                d = extract_date_from_filename(fname)
                rows = read_xlsx_first_sheet_rows(content)
                sold_val = parse_sold_jobs_last_estimate_id(rows)
                n = apply_7col_updates(wb, sheet_name, d, sold_jobs=sold_val)
                total_updates += n
                print(f"[{label} Sold Jobs] {fname} | file_date={d} | updated_cells={n}")

    handle_wo_and_sold(SHEET_BUILDER, msg_builder_wo, msg_builder_sold, "Builder")
    handle_wo_and_sold(SHEET_COMMERCIAL, msg_comm_wo, msg_comm_sold, "Commercial")
    handle_wo_and_sold(SHEET_MULTIFAMILY, msg_mf_wo, msg_mf_sold, "MultiFamily")

    # ================= Follow Up counts => Opportunities/Estimates =================
    if msg_followup:
        mid = msg_followup["id"]
        fname, content = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, mid)
        if content:
            d = extract_date_from_filename(fname)
            rows = read_xlsx_first_sheet_rows(content)

            b_cnt = parse_followup_counts_by_bu(rows, FOLLOWUP_BUILDER_BU)
            c_cnt = parse_followup_counts_by_bu(rows, FOLLOWUP_COMM_BU)
            m_cnt = parse_followup_counts_by_bu(rows, FOLLOWUP_MF_BU)

            n1 = apply_7col_updates(wb, SHEET_BUILDER, d, opp_est=b_cnt)
            n2 = apply_7col_updates(wb, SHEET_COMMERCIAL, d, opp_est=c_cnt)
            n3 = apply_7col_updates(wb, SHEET_MULTIFAMILY, d, opp_est=m_cnt)

            total_updates += (n1 + n2 + n3)
            print(f"[Follow Up Opp/Est] {fname} | file_date={d} | builder={b_cnt} comm={c_cnt} mf={m_cnt} | updated_cells={n1+n2+n3}")

    # ---- Upload once ----
    if total_updates > 0:
        out = io.BytesIO()
        wb.save(out)
        upload_sharepoint_excel(token, DRIVE_ID, ITEM_ID, out.getvalue())
        print(f"Done. Uploaded updated Excel back to SharePoint. total_updates={total_updates}")
    else:
        print("No cells updated. Nothing uploaded.")


if __name__ == "__main__":
    main()
